import os 
import random 
import time
import json
import torch
import torchvision
import numpy as np 
import warnings
import argparse



import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx write data successfully！")
 
 
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)

    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()



#4. more details to build main function    
        

def main(train_mode, alpha, models_family):

    if models_family == 'vgg':
        model_names = ['vgg11_bn', 'vgg13_bn', 'vgg16_bn', 'vgg19_bn'] 
    if models_family == 'resnet':
        model_names = ['resnet18', 'resnet34', 'resnet50', 'resnet101', 'resnet152']
    train = []
    val = []
    train.append(model_names); val.append(model_names)
    models = []
    for idx,model_name in enumerate(model_names):
        best_model = torch.load("./checkpoints_region/{}/{}/_checkpoint.pth.tar".format(model_name,alpha))
        models.append(best_model)
    for i in range(100):
        result_train = []
        result_val = []

        for model_id in range(len(models)):
            print(i,model_id)
            result_train.append(models[model_id]['result_train'][i][1].cpu());result_val.append(models[model_id]['result_val'][i][1].cpu())
        train.append(result_train); val.append(result_val)
                    
        

    write_excel_xlsx(f'{train_mode}_{models_family}_train.xlsx', 'top1', train)    
    write_excel_xlsx(f'{train_mode}_{models_family}_val.xlsx', 'top1', val)                
if __name__ =="__main__":

    parser = argparse.ArgumentParser(description='Result of Trained ship class network')
    parser.add_argument('--train_mode', help='The way to train networks patterns', type=str, default='Vanilla')
    parser.add_argument('--alpha', help='Parameters for adjusting the highlighting level of key areas', type=str, default='00')
    parser.add_argument('--models_family', help='Networks Family', type=str, default='vgg')
    args = parser.parse_args()
    main(args.train_mode, args.alpha, args.models_family)





















